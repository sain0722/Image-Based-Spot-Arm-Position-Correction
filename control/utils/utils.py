import json
import os

import cv2
import numpy as np
import openpyxl
import pandas as pd
# from PyQt5 import QtGui
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QPixmap, QImage, QColor
from PyQt5.QtWidgets import QLabel
from bosdyn.api import geometry_pb2
from bosdyn.client import math_helpers
from openpyxl.styles import Border, Side, Alignment


# Power 상태에 따른 색상 및 배경색 딕셔너리
POWER_STYLES = {
    "ON": {"color": QColor("#2ECC71"), "background-color": QColor("#646464")},  # 밝은 녹색, 연한 회색 배경
    "POWERING_ON": {"color": QColor("#F1C40F"), "background-color": QColor("#646464")},  # 밝은 노란색, 연한 회색 배경
    "POWERING_OFF": {"color": QColor("#FFA500"), "background-color": QColor("#646464")},  # 오렌지색, 연한 회색 배경
    "OFF": {"color": QColor("#E74C3C"), "background-color": QColor("#646464")},  # 밝은 빨간색, 연한 회색 배경
}


def get_unique_filename(path, filename):
    # 파일명과 확장자 분리
    basename, ext = os.path.splitext(filename)

    counter = 1
    while os.path.exists(os.path.join(path, filename)):
        # 파일명이 이미 존재하면, 파일명 뒤에 숫자를 붙여서 새로운 파일명 생성
        filename = f"{basename}_{counter}{ext}"
        counter += 1

    return filename


def get_position_and_rotation_from_label(widget, label_name):
    pos_axis = ['x', 'y', 'z']
    rot_axis = ['x', 'y', 'z', 'w']
    position_widgets = [
        getattr(widget, f"lbl_{label_name}_pos_{axis}")
        for axis in pos_axis
    ]
    rotation_widgets = [
        getattr(widget, f"lbl_{label_name}_rot_{axis}")
        for axis in rot_axis
    ]

    position = {axis: float(label.text()) for label, axis in zip(position_widgets, pos_axis)}
    rotation = {axis: float(label.text()) for label, axis in zip(rotation_widgets, rot_axis)}

    return position, rotation


def set_position_and_rotation(widget, label_name, position, rotation):
    pos_axis = ['x', 'y', 'z']
    rot_axis = ['x', 'y', 'z', 'w']
    position_widgets = [
        getattr(widget, f"lbl_{label_name}_pos_{axis}")
        for axis in pos_axis
    ]
    rotation_widgets = [
        getattr(widget, f"lbl_{label_name}_rot_{axis}")
        for axis in rot_axis
    ]

    for label, axis in zip(position_widgets, pos_axis):
        label.setText(str(position.get(axis, '')))

    for label, axis in zip(rotation_widgets, rot_axis):
        label.setText(str(rotation.get(axis, '')))


def read_transformation_file(file_path):
    transformation = []
    with open(file_path, 'r') as f:
        content = f.readlines()
        for line in content:
            value = line.split(",")
            value[-1] = value[-1][:-1]
            transformation.append(value)

    t = np.array(transformation, dtype=float)
    return t


def convert_to_target_pose(arm_position_real):
    target_pose = {
        "x": arm_position_real["position"]["x"],
        "y": arm_position_real["position"]["y"],
        "z": arm_position_real["position"]["z"],
        "rotation": {
            "w": arm_position_real["rotation"]["w"],
            "x": arm_position_real["rotation"]["x"],
            "y": arm_position_real["rotation"]["y"],
            "z": arm_position_real["rotation"]["z"],
        },
    }
    return target_pose


def get_qimage(image):
    if image.shape[-1] == 3:
        height, width, colors = image.shape
        bytesPerLine = 3 * width
        image_format = QImage.Format_RGB888

        # composing image from image data
        image = QImage(image.data,
                       width,
                       height,
                       bytesPerLine,
                       image_format)

        image = image.rgbSwapped()

    else:
        # 데이터 타입을 uint8로 변환합니다.
        depth_data_uint8 = cv2.convertScaleAbs(image, alpha=(255.0 / image.max()))

        # Convert the image to grayscale
        # gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        height, width = depth_data_uint8.shape
        bytesPerLine = width
        image_format = QImage.Format_Grayscale8

        # composing image from image data
        image = QImage(depth_data_uint8.data,
                       width,
                       height,
                       bytesPerLine,
                       image_format)

    return image


def get_qpixmap_grayscale(image):
    # 데이터 타입을 uint8로 변환합니다.
    depth_data_uint8 = cv2.convertScaleAbs(image, alpha=(255.0 / image.max()))

    height, width = depth_data_uint8.shape
    bytes_per_line = width
    image_format = QImage.Format_Grayscale8

    # 이미지 데이터를 복사합니다.
    image_data = depth_data_uint8.copy().data

    qimage = QImage(image_data, width, height, bytes_per_line, image_format)
    qpixmap = QPixmap.fromImage(qimage)
    return qpixmap


def get_key_from_label(label: QLabel):
    key = ""
    if label.objectName() == "lblSrcImage":
        key = "src_origin"
    elif label.objectName() == "lblMiddleImage":
        key = "middle_origin"
    elif label.objectName() == "lblRightImage":
        key = "right_origin"
    return key


def set_pixmap(fname: str, label: QLabel, is_draw_center_line: bool = True):
    image = cv2.imread(fname)

    if is_draw_center_line:
        image = draw_center_line(image)

    qimage = get_qimage(image)
    pixmap = QPixmap.fromImage(qimage)

    h, w = label.height(), label.width()
    scaled_pixmap = pixmap.scaled(w - 6, h - 6, aspectRatioMode=Qt.KeepAspectRatio)
    label.setPixmap(scaled_pixmap)
    return image


def set_pixmap_image(image, label: QLabel, is_draw_center_line: bool = True):
    if is_draw_center_line:
        image = draw_center_line(image)

    qimage = get_qimage(image)
    pixmap = QPixmap.fromImage(qimage)

    h, w = label.height(), label.width()
    scaled_pixmap = pixmap.scaled(w - 6, h - 6, aspectRatioMode=Qt.KeepAspectRatio)
    label.setPixmap(scaled_pixmap)
    return image


def draw_center_line(img):
    h, w, _ = img.shape
    row_start_point = (0, int(h / 2))
    row_end_point = (w - 1, int(h / 2))

    col_start_point = (int(w / 2), 0)
    col_end_point = (int(w / 2), h - 1)
    print(img.shape)
    thickness = get_thickness(w, h)
    # thickness = 5

    cv2.line(img, row_start_point, row_end_point, (0, 0, 255), thickness)
    cv2.line(img, col_start_point, col_end_point, (0, 0, 255), thickness)

    return img


def _draw_center_line(img):
    height, width, channels = img.shape
    center_x = int(width / 2)
    center_y = int(height / 2)

    thickness = get_thickness(width, height)

    # Draw a line from top to bottom of the image
    cv2.line(img, (center_x, 0), (center_x, height), (0, 0, 255), thickness=thickness)
    cv2.line(img, (center_y, 0), (center_y, width), (0, 0, 255), thickness=thickness)

    return img


def get_thickness(w: int, h: int) -> int:
    # 640 * 480
    # 1280 * 720
    # 1920 * 1080
    # 3840 * 2160
    # 4096 * 2160
    thickness: int = 2

    if w >= 1280 or h >= 1280:
        thickness = 3

    if w >= 1920 or h >= 1920:
        thickness = 6

    if w >= 3600 or h >= 3600:
        thickness = 10

    # if imsize >= 1280 * 720:
    #     thickness = 4
    #
    # if imsize >= 1920 * 1080:
    #     thickness = 6
    #
    # if imsize >= 3840 * 2160:
    #     thickness = 10

    return thickness


def get_power_style(power):
    # Set the color and background color of lblPowerValue based on power state
    style = POWER_STYLES.get(power, {"color": QColor("black"), "background-color": QColor("#646464")})
    return style


def set_label_color(label, color, background_color):
    label.setStyleSheet("color: {0}; background-color: {1};".format(color.name(), background_color.name()))


def set_status_time_value(progress_bar, status, value, time_left):
    progress_bar.setValue(value)
    progress_bar.setFormat(f"{status} - %p% - {time_left}")
    progress_bar.setTextVisible(True)

    # 배터리 상태에 따라 다른 색상을 설정
    if status == "CHARGING":
        bar_color = "SpringGreen"
        text_color = "black"
    elif status == "DISCHARGING":
        if value > 50:
            bar_color = "SkyBlue"
            text_color = "white"
        else:
            bar_color = "Tomato"
            text_color = "white"
    else:
        bar_color = "LightGray"
        text_color = "black"

    # 스타일 시트 설정
    progress_bar.setStyleSheet("""
        QProgressBar
        {
            border: 2px solid grey;
            border-radius: 5px;
            text-align: center;
            color: %s;
        }

        QProgressBar::chunk
        {
            background-color: %s;
        }
        """ % (text_color, bar_color))


def se3pose_to_dict(pose_proto):
    position = {
        'x': round(pose_proto.position.x, 4),
        'y': round(pose_proto.position.y, 4),
        'z': round(pose_proto.position.z, 4)
    }

    rotation = {
        'x': round(pose_proto.rotation.x, 6),
        'y': round(pose_proto.rotation.y, 6),
        'z': round(pose_proto.rotation.z, 6),
        'w': round(pose_proto.rotation.w, 6)
    }

    return position, rotation


def dict_to_se3pose(dict_data):
    position = geometry_pb2.Vec3(x=dict_data['position']['x'],
                                 y=dict_data['position']['y'],
                                 z=dict_data['position']['z'])

    rotation = geometry_pb2.Quaternion(w=dict_data['rotation']['w'],
                                       x=dict_data['rotation']['x'],
                                       y=dict_data['rotation']['y'],
                                       z=dict_data['rotation']['z'])

    se3pose = math_helpers.SE3Pose(x=position.x,
                                   y=position.y,
                                   z=position.z,
                                   rot=rotation)

    return se3pose


def get_arm_position_dict(name, transform):
    arm_position_dict = {
        "name": name,
        "transform": transform
    }

    return arm_position_dict


def create_json_format(fid_id, dist_margin, data):
    json_data = {
        'fid_id': fid_id,
        'dist_margin': dist_margin,
        'frame_tform_gripper': data
    }
    return json_data


def create_json_file(saved_name, data):
    json_data = json.dumps(data, indent=4)  # 데이터를 JSON 형식으로 변환하고 들여쓰기 설정
    with open(saved_name, "w") as file:
        file.write(json_data)


def overlap_images(image1, image2):
    overlapped = cv2.addWeighted(image1, 0.5, image2, 0.5, 0)
    return overlapped


class ExcelManager:
    start_cell = 'B'
    end_cell = 'Y'

    @staticmethod
    def create(saved_name):
        # 엑셀 파일 열기
        wb = openpyxl.Workbook()

        # 시트 선택
        sheet = wb.active
        sheet.column_dimensions['B'].width = 16.25
        sheet.column_dimensions['N'].width = 16.25

        # 셀 병합
        sheet.merge_cells('B1:M1')
        sheet['B1'] = '첫번째 위치'

        sheet.merge_cells('N1:Y1')
        sheet['N1'] = '두번째 위치'

        # 셀 범위 선택
        cell_range = sheet['B1:Y1']

        # 가운데 정렬 스타일 설정
        center_alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

        # 테두리 스타일 설정
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        # 병합된 셀에 스타일 적용
        sheet['B1'].alignment = center_alignment
        sheet['N1'].alignment = center_alignment

        # 셀 테두리 적용
        for row in cell_range:
            for cell in row:
                cell.border = thin_border

        # 변경 내용 저장
        wb.save(saved_name)

        # 시트 선택
        writer = pd.ExcelWriter(saved_name, engine='openpyxl')
        writer.book = wb
        writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)

        columns = ['time', 'input_x', 'input_y', 'input_z', 'sh0', 'sh1', 'el0', 'el1', 'wr0', 'wr1', 'hand_position', 'hand_rotation',
                   'time', 'input_x', 'input_y', 'input_z', 'sh0', 'sh1', 'el0', 'el1', 'wr0', 'wr1', 'hand_position', 'hand_rotation']

        # 데이터프레임 생성
        df = pd.DataFrame(columns=columns)

        # 데이터프레임과 함께 헤더 추가
        df.to_excel(writer, sheet_name='Sheet', index=False, startrow=1, startcol=1, header=True)
        # worksheet = sheet.sheets['Sheet']
        writer.save()

    @staticmethod
    def append_data_to_excel(df, file_name, func):
        # 엑셀 파일 불러오기
        wb = openpyxl.load_workbook(file_name)

        # 시트 선택
        sheet = wb.active
        writer = pd.ExcelWriter(file_name, engine='openpyxl')
        writer.book = wb
        writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)

        # 시트의 행 개수 가져오기
        row_count = sheet.max_row

        # 테두리 스타일 설정
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        # 가운데 정렬 스타일 설정
        center_alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

        # 셀 범위 선택
        start_cell = 'B' + str(row_count + 1)
        end_cell   = 'Y' + str(row_count + 1)
        cell_range = sheet[f'{start_cell}:{end_cell}']
        # 셀 테두리 적용
        for row in cell_range:
            for cell in row:
                cell.border = thin_border
                cell.alignment = center_alignment

        # 데이터프레임 엑셀 파일에 추가
        df.to_excel(writer, sheet_name='Sheet', index=False, startrow=row_count, startcol=1, header=False)

        # 변경 내용 저장
        try:
            wb.save(file_name)
        except PermissionError:
            message = "엑셀 파일이 열려있는지 확인하세요."
            print(message)
            func.show_message_box(message)
        except Exception as e:
            print(e)

